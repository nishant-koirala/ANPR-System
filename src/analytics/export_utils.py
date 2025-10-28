"""
Export utilities for analytics reports
Supports PDF and Excel export with charts
"""

import os
from datetime import datetime
from typing import Dict, List, Optional
import io

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import matplotlib
    matplotlib.use('Agg')  # Use non-GUI backend
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False


class ReportExporter:
    """Export analytics reports to PDF and Excel formats"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    # ===== PDF Export =====
    
    def export_to_pdf(self, report_data: Dict, filename: Optional[str] = None) -> str:
        """Export report to PDF with charts"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"anpr_report_{timestamp}.pdf"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("ANPR System Analytics Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Report metadata
        meta_style = styles['Normal']
        story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
        story.append(Paragraph(f"<b>Report Period:</b> {report_data.get('period', 'N/A')}", meta_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Summary Statistics
        if 'summary' in report_data:
            story.append(Paragraph("<b>Summary Statistics</b>", styles['Heading2']))
            summary = report_data['summary']
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Entries', str(summary.get('total_entries', 0))],
                ['Total Exits', str(summary.get('total_exits', 0))],
                ['Current Occupancy', str(summary.get('current_occupancy', 0))],
                ['Total Revenue', f"NPR {summary.get('total_revenue', 0):.2f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Revenue Analysis
        if 'revenue' in report_data:
            story.append(Paragraph("<b>Revenue Analysis</b>", styles['Heading2']))
            revenue = report_data['revenue']
            
            revenue_data = [
                ['Metric', 'Value'],
                ['Total Revenue', f"NPR {revenue.get('total_revenue', 0):.2f}"],
                ['Average Revenue', f"NPR {revenue.get('avg_revenue', 0):.2f}"],
                ['Maximum Revenue', f"NPR {revenue.get('max_revenue', 0):.2f}"],
                ['Minimum Revenue', f"NPR {revenue.get('min_revenue', 0):.2f}"],
                ['Total Hours', f"{revenue.get('total_hours', 0):.2f} hrs"]
            ]
            
            revenue_table = Table(revenue_data, colWidths=[3*inch, 2*inch])
            revenue_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(revenue_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Peak Hours
        if 'peak_hours' in report_data:
            story.append(Paragraph("<b>Peak Hours</b>", styles['Heading2']))
            peak_hours = report_data['peak_hours']
            
            if peak_hours:
                peak_data = [['Hour', 'Entries', 'Exits', 'Total']]
                for peak in peak_hours[:5]:  # Top 5 peak hours
                    hour = peak['hour']
                    data = peak['data']
                    peak_data.append([
                        f"{hour}:00",
                        str(data['entries']),
                        str(data['exits']),
                        str(data['total'])
                    ])
                
                peak_table = Table(peak_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
                peak_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(peak_table)
                story.append(Spacer(1, 0.3*inch))
        
        # Patterns
        if 'patterns' in report_data:
            story.append(Paragraph("<b>Usage Patterns</b>", styles['Heading2']))
            patterns = report_data['patterns']
            
            pattern_data = [
                ['Metric', 'Value'],
                ['Average Duration', f"{patterns.get('avg_duration_hours', 0):.2f} hours"],
                ['Peak Day', patterns.get('peak_day', {}).get('day', 'N/A')],
                ['Exit Rate', f"{patterns.get('exit_rate', 0):.2f}%"],
                ['Current Occupancy', str(patterns.get('current_occupancy', 0))]
            ]
            
            pattern_table = Table(pattern_data, colWidths=[3*inch, 2*inch])
            pattern_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(pattern_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Forecast
        if 'forecast' in report_data:
            story.append(PageBreak())
            story.append(Paragraph("<b>Demand Forecast (Next 7 Days)</b>", styles['Heading2']))
            forecast = report_data['forecast']
            
            if forecast.get('forecast'):
                forecast_data = [['Date', 'Predicted Entries', 'Confidence']]
                for item in forecast['forecast']:
                    forecast_data.append([
                        item['date'],
                        str(item['predicted_entries']),
                        item['confidence'].capitalize()
                    ])
                
                forecast_table = Table(forecast_data, colWidths=[2*inch, 2*inch, 2*inch])
                forecast_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(forecast_table)
        
        # Build PDF
        doc.build(story)
        return filepath
    
    # ===== Excel Export =====
    
    def export_to_excel(self, report_data: Dict, filename: Optional[str] = None) -> str:
        """Export report to Excel with charts"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"anpr_report_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Header styles
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary['A1'] = "ANPR System Analytics Report"
        ws_summary['A1'].font = Font(bold=True, size=16, color="1a237e")
        ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        if 'summary' in report_data:
            summary = report_data['summary']
            ws_summary['A4'] = "Summary Statistics"
            ws_summary['A4'].font = Font(bold=True, size=14)
            
            headers = ['Metric', 'Value']
            for col, header in enumerate(headers, start=1):
                cell = ws_summary.cell(row=5, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            metrics = [
                ('Total Entries', summary.get('total_entries', 0)),
                ('Total Exits', summary.get('total_exits', 0)),
                ('Current Occupancy', summary.get('current_occupancy', 0)),
                ('Total Revenue (NPR)', f"{summary.get('total_revenue', 0):.2f}")
            ]
            
            for row, (metric, value) in enumerate(metrics, start=6):
                ws_summary.cell(row=row, column=1, value=metric).border = border
                ws_summary.cell(row=row, column=2, value=value).border = border
            
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20
        
        # Revenue Sheet
        if 'revenue' in report_data:
            ws_revenue = wb.create_sheet("Revenue Analysis")
            ws_revenue['A1'] = "Revenue Analysis"
            ws_revenue['A1'].font = Font(bold=True, size=14)
            
            revenue = report_data['revenue']
            
            headers = ['Metric', 'Value']
            for col, header in enumerate(headers, start=1):
                cell = ws_revenue.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            metrics = [
                ('Total Revenue (NPR)', f"{revenue.get('total_revenue', 0):.2f}"),
                ('Average Revenue (NPR)', f"{revenue.get('avg_revenue', 0):.2f}"),
                ('Maximum Revenue (NPR)', f"{revenue.get('max_revenue', 0):.2f}"),
                ('Minimum Revenue (NPR)', f"{revenue.get('min_revenue', 0):.2f}"),
                ('Total Hours', f"{revenue.get('total_hours', 0):.2f}")
            ]
            
            for row, (metric, value) in enumerate(metrics, start=4):
                ws_revenue.cell(row=row, column=1, value=metric).border = border
                ws_revenue.cell(row=row, column=2, value=value).border = border
            
            ws_revenue.column_dimensions['A'].width = 25
            ws_revenue.column_dimensions['B'].width = 20
        
        # Daily Trends Sheet
        if 'daily_trends' in report_data:
            ws_trends = wb.create_sheet("Daily Trends")
            ws_trends['A1'] = "Daily Entry/Exit Trends"
            ws_trends['A1'].font = Font(bold=True, size=14)
            
            headers = ['Date', 'Entries', 'Exits']
            for col, header in enumerate(headers, start=1):
                cell = ws_trends.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            trends = report_data['daily_trends']
            for row, (date, data) in enumerate(sorted(trends.items()), start=4):
                ws_trends.cell(row=row, column=1, value=date).border = border
                ws_trends.cell(row=row, column=2, value=data['entries']).border = border
                ws_trends.cell(row=row, column=3, value=data['exits']).border = border
            
            # Add chart
            chart = LineChart()
            chart.title = "Daily Trends"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Date"
            
            data_ref = Reference(ws_trends, min_col=2, min_row=3, max_col=3, max_row=3+len(trends))
            cats_ref = Reference(ws_trends, min_col=1, min_row=4, max_row=3+len(trends))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            ws_trends.add_chart(chart, "E3")
            
            ws_trends.column_dimensions['A'].width = 15
            ws_trends.column_dimensions['B'].width = 12
            ws_trends.column_dimensions['C'].width = 12
        
        # Peak Hours Sheet
        if 'peak_hours' in report_data:
            ws_peaks = wb.create_sheet("Peak Hours")
            ws_peaks['A1'] = "Peak Hours Analysis"
            ws_peaks['A1'].font = Font(bold=True, size=14)
            
            headers = ['Hour', 'Entries', 'Exits', 'Total']
            for col, header in enumerate(headers, start=1):
                cell = ws_peaks.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            peak_hours = report_data['peak_hours']
            for row, peak in enumerate(peak_hours, start=4):
                hour = peak['hour']
                data = peak['data']
                ws_peaks.cell(row=row, column=1, value=f"{hour}:00").border = border
                ws_peaks.cell(row=row, column=2, value=data['entries']).border = border
                ws_peaks.cell(row=row, column=3, value=data['exits']).border = border
                ws_peaks.cell(row=row, column=4, value=data['total']).border = border
            
            ws_peaks.column_dimensions['A'].width = 12
            ws_peaks.column_dimensions['B'].width = 12
            ws_peaks.column_dimensions['C'].width = 12
            ws_peaks.column_dimensions['D'].width = 12
        
        # Forecast Sheet
        if 'forecast' in report_data and report_data['forecast'].get('forecast'):
            ws_forecast = wb.create_sheet("Forecast")
            ws_forecast['A1'] = "Demand Forecast (Next 7 Days)"
            ws_forecast['A1'].font = Font(bold=True, size=14)
            
            headers = ['Date', 'Predicted Entries', 'Confidence']
            for col, header in enumerate(headers, start=1):
                cell = ws_forecast.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            forecast = report_data['forecast']['forecast']
            for row, item in enumerate(forecast, start=4):
                ws_forecast.cell(row=row, column=1, value=item['date']).border = border
                ws_forecast.cell(row=row, column=2, value=item['predicted_entries']).border = border
                ws_forecast.cell(row=row, column=3, value=item['confidence'].capitalize()).border = border
            
            ws_forecast.column_dimensions['A'].width = 15
            ws_forecast.column_dimensions['B'].width = 18
            ws_forecast.column_dimensions['C'].width = 15
        
        # Save workbook
        wb.save(filepath)
        return filepath
    
    # ===== Chart Generation =====
    
    def generate_chart_image(self, chart_type: str, data: Dict, filename: str) -> Optional[str]:
        """Generate chart image using matplotlib"""
        if not MATPLOTLIB_AVAILABLE:
            return None
        
        filepath = os.path.join(self.output_dir, filename)
        
        plt.figure(figsize=(10, 6))
        
        if chart_type == 'line':
            dates = list(data.keys())
            values = list(data.values())
            plt.plot(dates, values, marker='o', linewidth=2)
            plt.xticks(rotation=45)
            plt.grid(True, alpha=0.3)
        
        elif chart_type == 'bar':
            labels = list(data.keys())
            values = list(data.values())
            plt.bar(labels, values, color='#1a237e')
            plt.xticks(rotation=45)
        
        elif chart_type == 'pie':
            labels = list(data.keys())
            values = list(data.values())
            plt.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
        
        plt.tight_layout()
        plt.savefig(filepath, dpi=150, bbox_inches='tight')
        plt.close()
        
        return filepath
